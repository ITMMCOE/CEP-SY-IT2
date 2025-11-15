from datetime import date, timedelta
from statistics import mean
from django.db.models import Sum, Max
from django.utils import timezone
from .models import InventoryBatch, Product, ProductDailySales, SupplierProduct, StockAlert, ProductStockSnapshot
import csv
from pathlib import Path

try:
    import openpyxl  # for Excel parsing
    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    openpyxl = None
    OPENPYXL_AVAILABLE = False

def remove_stock_fifo(product, quantity):
    """
    Removes stock from batches for the given product using FIFO logic.
    Returns True if enough stock was available and removed, False otherwise.
    """
    batches = InventoryBatch.objects.filter(product=product, quantity__gt=0).order_by('received_at')
    remaining = quantity
    for batch in batches:
        if batch.quantity >= remaining:
            batch.quantity -= remaining
            batch.save()
            remaining = 0
            break
        else:
            remaining -= batch.quantity
            batch.quantity = 0
            batch.save()
    if remaining == 0:
        product.current_stock -= quantity
        product.save()
        return True
    else:
        # Not enough stock!
        return False


# --- Reorder & Safety Stock Calculations ---

def get_daily_sales_window(product: Product, days: int = 90):
    """Return list of last N days sales quantities (0 for missing days)."""
    end = date.today()
    start = end - timedelta(days=days - 1)
    qs = ProductDailySales.objects.filter(product=product, date__range=(start, end))
    # Map by date for quick lookup
    by_date = {row.date: row.quantity for row in qs}
    quantities = []
    d = start
    while d <= end:
        quantities.append(by_date.get(d, 0))
        d += timedelta(days=1)
    return quantities


def average_daily_usage(product: Product, days: int = 30):
    data = get_daily_sales_window(product, days)
    if not data:
        return 0
    return sum(data) / len(data)


def maximum_daily_sales(product: Product, days: int = 90):
    data = get_daily_sales_window(product, days)
    return max(data) if data else 0


def lead_time_stats(product: Product):
    """Return (average_lead_time, maximum_lead_time) in days from SupplierProduct entries."""
    qs = SupplierProduct.objects.filter(product=product, lead_time_days__isnull=False)
    times = list(qs.values_list('lead_time_days', flat=True))
    if not times:
        return (0, 0)
    avg = sum(times) / len(times)
    return (avg, max(times))


def safety_stock_basic(product: Product):
    """Basic safety stock: use difference of variability; fallback simple rule."""
    # We'll approximate with (max daily - avg daily) * average lead time
    avg_usage = average_daily_usage(product)
    max_usage = maximum_daily_sales(product)
    avg_lt, _ = lead_time_stats(product)
    return max(0, (max_usage - avg_usage) * avg_lt)


def safety_stock_advanced(product: Product):
    """Advanced formula user provided:
    Safety Stock = (Maximum Daily Sales × Maximum Lead Time) - (Average Daily Sales × Average Lead Time)
    """
    avg_usage = average_daily_usage(product)
    max_usage = maximum_daily_sales(product)
    avg_lt, max_lt = lead_time_stats(product)
    return max(0, (max_usage * max_lt) - (avg_usage * avg_lt))


def reorder_point(product: Product, use_advanced: bool = True):
    avg_usage = average_daily_usage(product)
    avg_lt, _ = lead_time_stats(product)
    safety = safety_stock_advanced(product) if use_advanced else safety_stock_basic(product)
    return int(round((avg_usage * avg_lt) + safety))


def reorder_recommendation(product: Product):
    point = reorder_point(product)
    current = product.current_stock
    delta = point - current
    needs_reorder = delta > 0
    preferred_supplier = SupplierProduct.objects.filter(product=product, is_preferred=True).order_by('cost_price').first()
    supplier_info = None
    if preferred_supplier:
        supplier_info = {
            'supplier_id': preferred_supplier.supplier_id,
            'supplier_name': preferred_supplier.supplier.name,
            'lead_time_days': preferred_supplier.lead_time_days,
            'cost_price': float(preferred_supplier.cost_price) if preferred_supplier.cost_price is not None else None,
            'min_order_quantity': preferred_supplier.min_order_quantity,
        }
    return {
        'product_id': product.id,
        'product_name': product.name,
        'sku': product.sku,
        'current_stock': current,
        'reorder_point': point,
        'recommended_order_quantity': max(delta, 0),
        'safety_stock_basic': int(round(safety_stock_basic(product))),
        'safety_stock_advanced': int(round(safety_stock_advanced(product))),
        'average_daily_usage': average_daily_usage(product),
        'maximum_daily_sales': maximum_daily_sales(product),
        'needs_reorder': needs_reorder,
        'preferred_supplier': supplier_info,
    }


def all_reorder_recommendations():
    return [reorder_recommendation(p) for p in Product.objects.all().order_by('name')]


# --- Simple Alert Evaluation (minimum stock + buffer) ---

def evaluate_product_alert(product: Product, save: bool = True):
    """Evaluate product's reorder_status & manage StockAlert records.

    Status logic:
        LOW: current_stock <= minimum_stock_level
        APPROACHING: current_stock <= minimum_stock_level * (1 + buffer_pct/100) but > minimum
        OK: else
    Creates a StockAlert when entering APPROACHING or LOW, resolves active alerts when status returns to OK.
    """
    minimum = max(product.minimum_stock_level, 0)
    buffer_multiplier = 1 + (product.reorder_warning_buffer_pct / 100.0)
    approaching_threshold = int(minimum * buffer_multiplier)

    current = product.current_stock
    if current <= minimum:
        new_status = Product.STATUS_LOW
    elif current <= approaching_threshold and minimum > 0:
        new_status = Product.STATUS_APPROACHING
    else:
        new_status = Product.STATUS_OK

    # Update product status/time
    if new_status != product.reorder_status:
        product.reorder_status = new_status
        product.reorder_status_changed_at = timezone.now()
        if save:
            product.save(update_fields=['reorder_status', 'reorder_status_changed_at'])

    # Resolve active alerts if OK
    if new_status == Product.STATUS_OK:
        StockAlert.objects.filter(product=product, active=True).update(active=False, resolved_at=timezone.now())
        return new_status

    # If APPROACHING or LOW ensure an active alert exists (one per status at a time)
    existing = StockAlert.objects.filter(product=product, status=new_status, active=True).first()
    if not existing:
        message = (
            f"Stock {'below' if new_status == Product.STATUS_LOW else 'approaching'} minimum. "
            f"Current={current}, Minimum={minimum} (Buffer {product.reorder_warning_buffer_pct}%)."
        )
        StockAlert.objects.create(
            product=product,
            status=new_status,
            current_stock_at_trigger=current,
            minimum_stock_level=minimum,
            message=message,
        )
    return new_status


def evaluate_all_alerts():
    results = []
    for p in Product.objects.all():
        status = evaluate_product_alert(p, save=True)
        results.append({'product_id': p.id, 'sku': p.sku, 'status': status, 'current_stock': p.current_stock})
    return results


def create_daily_stock_snapshots(date=None):
    """Create (or skip existing) stock snapshots for all products for a given date."""
    from datetime import date as date_cls
    if date is None:
        date = date_cls.today()
    created = 0
    for p in Product.objects.all():
        obj, was_created = ProductStockSnapshot.objects.get_or_create(
            product=p,
            date=date,
            defaults={'stock_level': p.current_stock}
        )
        if was_created:
            created += 1
    return {'date': str(date), 'created': created}


# --- Importers: Excel -> CSV -> DB ---

REQUIRED_HEADERS = [
    'SR NO.', 'PRODUCT NAME', 'PRODUCT ID', 'PRODUCTS ORDERED',
    'PRODUCTS USED', 'PRODUCTS IN STOCK', 'RESTOCK'
]


def convert_excel_to_csv(input_path, output_csv_path=None):
    """Convert first worksheet of an Excel file to CSV matching REQUIRED_HEADERS.

    Returns path to generated CSV. Raises ValueError if headers missing.
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl not installed. Please install openpyxl.")

    input_path = Path(input_path)
    wb = openpyxl.load_workbook(filename=str(input_path), data_only=True)
    ws = wb.active

    # Read header row
    headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # Basic normalization: upper-case compare
    normalized = [h.upper() for h in headers]
    required_upper = [h.upper() for h in REQUIRED_HEADERS]
    if normalized[:len(required_upper)] != required_upper:
        raise ValueError(f"Excel headers do not match required format. Found: {headers}")

    # Prepare CSV path
    if output_csv_path is None:
        output_csv_path = input_path.with_suffix('.csv')
    output_csv_path = Path(output_csv_path)

    with output_csv_path.open('w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(REQUIRED_HEADERS)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue
            writer.writerow(list(row[:len(REQUIRED_HEADERS)]))

    return str(output_csv_path)


def purge_all_inventory_data():
    """Dangerous: delete all inventory-related data so a fresh import becomes the new source of truth.

    Deletes in order to satisfy FKs. Suppliers are preserved.
    """
    # Delete dependent/history first
    ProductDailySales.objects.all().delete()
    StockAlert.objects.all().delete()
    ProductStockSnapshot.objects.all().delete()
    InventoryBatch.objects.all().delete()
    # Finally delete products (cascades to SupplierProduct via FK)
    Product.objects.all().delete()


def import_inventory_csv(csv_path, mode: str = 'append'):
    """Import inventory from a CSV file.

    mode:
      - 'append' (default): upsert products and create batches; keeps existing data.
      - 'replace_all': Purge all products/history, then import fresh.

    Returns: {'created': int, 'updated': int, 'mode': str}
    """
    csv_path = Path(csv_path)
    if not csv_path.exists():
        raise FileNotFoundError(csv_path)

    created = 0
    updated = 0
    if mode == 'replace_all':
        purge_all_inventory_data()
    with csv_path.open('r', newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        # Header check
        found_headers = [h.strip() for h in reader.fieldnames]
        print(f"DEBUG: Found headers: {found_headers}")
        print(f"DEBUG: Required headers: {REQUIRED_HEADERS}")
        
        if found_headers != REQUIRED_HEADERS:
            raise ValueError(f"CSV headers do not match required format. Found: {found_headers}, Expected: {REQUIRED_HEADERS}")

        for row in reader:
            sku = str(row['PRODUCT ID']).strip()
            name = str(row['PRODUCT NAME']).strip()
            print(f"DEBUG: Processing SKU: {sku}, Name: {name}")
            try:
                current_stock = int(float(row['PRODUCTS IN STOCK']))
            except Exception:
                current_stock = 0
            restock_text = str(row.get('RESTOCK', '')).upper()
            min_stock = 15 if 'NEEDS' in restock_text else 5

            product, was_created = Product.objects.update_or_create(
                sku=sku,
                defaults={
                    'name': name,
                    'minimum_stock_level': max(min_stock, 0),
                    'current_stock': max(current_stock, 0),
                }
            )
            print(f"DEBUG: Product {sku} - Created: {was_created}")

            print(f"DEBUG: Product {sku} - Created: {was_created}")

            if was_created:
                created += 1
            else:
                updated += 1

            if current_stock > 0:
                InventoryBatch.objects.create(
                    product=product,
                    quantity=current_stock,
                    received_at=timezone.now()
                )

            # Evaluate alerts per product post-import
            try:
                evaluate_product_alert(product, save=True)
            except Exception:
                pass

    print(f"DEBUG: Final counts - Created: {created}, Updated: {updated}")
    return {'created': created, 'updated': updated, 'mode': mode}